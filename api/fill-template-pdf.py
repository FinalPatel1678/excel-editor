import os
import json
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook
from fpdf import FPDF
from io import BytesIO

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../templates")
TEMPLATE_DIR = os.path.abspath(TEMPLATE_DIR)

# Define the path to your fonts (ensure you place the fonts in the correct location)
font_path_english = os.path.join(os.path.dirname(__file__), "../fonts/NotoSans-Regular.ttf")
font_path_hindi = os.path.join(os.path.dirname(__file__), "../fonts/NotoSansDevanagari-Regular.ttf")
font_path_gujarati = os.path.join(os.path.dirname(__file__), "../fonts/NotoSansGujarati-Regular.ttf")

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Parse incoming request body
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            request_data = json.loads(post_data.decode('utf-8'))

            file_name = request_data.get('fileName')
            form_data = request_data.get('formData')

            if not file_name or not form_data:
                self.send_error(400, "File name and form data are required")
                return

            file_path = os.path.join(TEMPLATE_DIR, file_name)

            if not os.path.exists(file_path):
                self.send_error(404, f"File not found: {file_path}")
                return

            # Load the workbook
            workbook = load_workbook(file_path, data_only=True)  # `data_only=True` fetches the evaluated results of formulas

            # Process workbook by replacing placeholders with form data
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and "{{" in cell.value and "}}" in cell.value:
                            placeholder = cell.value.strip("{{").strip("}}").strip()
                            if placeholder in form_data:
                                cell.value = form_data[placeholder]

            # Generate PDF
            pdf_buffer = self.generate_pdf(workbook, file_name)

            # Send the PDF response
            self.send_response(200)
            self.send_header("Content-Disposition", f"attachment; filename=filled_{file_name.replace('.xlsx', '.pdf')}")
            self.send_header("Content-Type", "application/pdf")
            self.end_headers()

            # Write the buffer content to the response
            pdf_buffer.seek(0)  # Ensure the buffer is at the beginning
            self.wfile.write(pdf_buffer.read())

        except Exception as e:
            import traceback
            print(f"Error processing request: {str(e)}")
            print("Traceback:", traceback.format_exc())
            self.send_error(500, f"Internal server error: {str(e)}")

    def generate_pdf(self, workbook, file_name):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Add fonts for English, Hindi, and Gujarati
        pdf.add_font("NotoSans", "", font_path_english, uni=True)  # English
        pdf.add_font("NotoSansHindi", "", font_path_hindi, uni=True)  # Hindi
        pdf.add_font("NotoSansGujarati", "", font_path_gujarati, uni=True)  # Gujarati

        # Set default font to English (NotoSans)
        pdf.set_font("NotoSans", size=12)

        # Iterate over the sheets in the workbook
        for sheet in workbook.worksheets:
            content_added = False  # Flag to track if content was added to the page
            
            # Initialize page flag and start the content tracking
            row_height = 10  # Height of a row in the PDF
            current_y = 20  # Initial position for the first row
            max_width = 0  # Track the maximum width of any cell (to handle column width)
            content_height = 0  # Track the overall height of content for the sheet

            # Iterate over the rows and columns in the sheet
            for row in sheet.iter_rows():
                row_text = ""
                for col_idx, cell in enumerate(row):
                    if cell.value is None or cell.value == "":
                        continue  # Skip empty cells

                    # Get cell value and font size (adjust based on language)
                    cell_value = str(cell.value)

                    # Set the appropriate font for different languages
                    if any("\u0900" <= char <= "\u097F" for char in cell_value):  # Hindi Unicode Range
                        pdf.set_font("NotoSansHindi", size=12)
                    elif any("\u0A80" <= char <= "\u0AFF" for char in cell_value):  # Gujarati Unicode Range
                        pdf.set_font("NotoSansGujarati", size=12)
                    else:  # Default to English
                        pdf.set_font("NotoSans", size=12)

                    # Accumulate text to row
                    row_text += cell_value + " "

                    # Calculate column width based on the longest cell content
                    max_width = max(max_width, pdf.get_string_width(cell_value))

                # Only add content to the PDF if there's text in the row
                if row_text.strip():
                    if not content_added:
                        pdf.add_page()  # Add a new page only if content will be written
                        content_added = True

                    # Write row content with proper spacing (acting like a paragraph)
                    pdf.set_xy(10, current_y)  # Set the position for this row
                    pdf.multi_cell(max_width, row_height + 4, row_text.strip(), align='L')

                    # Move the Y position for the next row
                    current_y += row_height + 4  # Add more space between rows for paragraph-like appearance

                    # Check if the current Y position exceeds the page height
                    if current_y + row_height + 4 > 270:  # 270 is the default page height in FPDF
                        pdf.add_page()
                        current_y = 20  # Reset to the top of the new page

        # Output the PDF to a BytesIO buffer
        pdf_buffer = BytesIO()
        pdf_output = pdf.output(dest='S').encode('latin1')  # Output PDF as string and encode it to bytes
        pdf_buffer.write(pdf_output)
        pdf_buffer.seek(0)  # Rewind the buffer to the beginning

        return pdf_buffer
