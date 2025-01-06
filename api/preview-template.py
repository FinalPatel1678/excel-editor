import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from http.server import BaseHTTPRequestHandler
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../templates")
TEMPLATE_DIR = os.path.abspath(TEMPLATE_DIR)

def serialize_datetime(obj):
    if isinstance(obj, datetime) and not pd.isna(obj):  # Check if it's a valid datetime and not NaT
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    return None  # Return None for NaT or any non-datetime objects

def extract_cell_styles(sheet):
    """Extract cell styles including background color, borders, font, alignment, row height, column width, merged cells, hyperlinks, and more."""
    cell_styles = {}
    
    # Extract row heights
    row_heights = {}
    for row in sheet.iter_rows():
        row_heights[row[0].row] = sheet.row_dimensions[row[0].row].height
    
    # Extract column widths
    column_widths = {}
    for col in sheet.columns:
        column_letter = get_column_letter(col[0].column)  # Convert column index to letter
        column_widths[column_letter] = sheet.column_dimensions[column_letter].width
    
    # Extract merged cells
    merged_cells = []
    for merge in sheet.merged_cells:
        merged_cells.append(merge)
    
    for row in sheet.iter_rows():
        for cell in row:
            bg_color = None
            if cell.fill.start_color.index != '00000000':  # Check if color is not transparent
                if cell.fill.start_color.index == 'FFFFFFFF':  # White color check
                    bg_color = '#FFFFFF'  # Use standard white color
                else:
                    bg_color = f"#{cell.fill.start_color.rgb[2:]}"  # Extract RGB color code
            
            font = {
                'color': cell.font.color.rgb if cell.font.color else None,  # Font color (RGB)
                'size': cell.font.size if cell.font.size else None,  # Font size
                'bold': cell.font.bold,  # Bold
                'italic': cell.font.italic,  # Italic
                'underline': cell.font.underline,  # Underline
            }
            
            border = {
                'top': {
                    'style': cell.border.top.style,
                    'color': cell.border.top.color.rgb if cell.border.top.color else None,
                },
                'right': {
                    'style': cell.border.right.style,
                    'color': cell.border.right.color.rgb if cell.border.right.color else None,
                },
                'bottom': {
                    'style': cell.border.bottom.style,
                    'color': cell.border.bottom.color.rgb if cell.border.bottom.color else None,
                },
                'left': {
                    'style': cell.border.left.style,
                    'color': cell.border.left.color.rgb if cell.border.left.color else None,
                }
            }
            cell_styles[cell.coordinate] = {
                'bg_color': bg_color,
                'font': font,
                'border': border,
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical
                },
                'row_height': row_heights.get(cell.row, None),  # row height if available
                'col_width': column_widths.get(get_column_letter(cell.column), None),  # column width if available
                'merged': cell.coordinate in [cell.coordinate for merge in merged_cells for cell in merge],
                'hyperlink': cell.hyperlink.target if cell.hyperlink else None,
                'number_format': cell.number_format if cell.number_format else None,
                'formula': cell.formula if hasattr(cell, 'formula') else None,  # Safely access formula attribute
                'comment': cell.comment.text if cell.comment else None
            }
    
    return cell_styles

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Parse incoming request body
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            request_data = json.loads(post_data.decode('utf-8'))
            file_name = request_data.get('fileName')
            form_data = request_data.get('formData')
            
            if not file_name:
                self.send_error(400, "File name is required")
                return
                
            file_path = os.path.join(TEMPLATE_DIR, file_name)
            
            if not os.path.exists(file_path):
                self.send_error(404, f"File not found: {file_path}")
                return
            
            # Load workbook and extract sheets and styles
            wb = load_workbook(file_path)
            all_sheets_data = {}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Extract styles for each sheet
                cell_styles = extract_cell_styles(sheet)

                # Apply form data to placeholders
                for column in df.columns:
                    df[column] = df[column].apply(lambda x: 
                        form_data.get(x[2:-2].strip(), x) if isinstance(x, str) and '{{' in x and '}}' in x  
                        else x
                    )

                # Convert sheet data to list of lists with styles
                sheet_data = []
                headers = df.columns.tolist()
                sheet_data.append(headers)

                for _, row in df.iterrows():
                    row_data = []
                    row_styles = []
                    for idx, value in enumerate(row):
                        # Handle NaN values (convert to None or another placeholder)
                        if pd.isna(value):
                            value = None  # Replace NaN with None

                        # Apply styles to each cell
                        cell_key = f"{chr(65 + idx)}{_ + 2}"  # Get cell name (e.g., 'A2')
                        style = cell_styles.get(cell_key, {})
                        row_data.append({
                            'value': value,
                            'styles': style
                        })
                    sheet_data.append(row_data)

                all_sheets_data[sheet_name] = sheet_data
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(all_sheets_data, default=str).encode('utf-8'))
                
        except Exception as e:
            print(f"Error processing request: {str(e)}")
            self.send_error(500, f"Internal server error: {str(e)}")
