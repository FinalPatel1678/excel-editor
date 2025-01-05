import os
import json
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook
from urllib.parse import parse_qs, urlparse

# Load the template directory path from a config (similar to the config.js in Node.js)
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../templates")
TEMPLATE_DIR = os.path.abspath(TEMPLATE_DIR)

class Handler(BaseHTTPRequestHandler):

    def do_GET(self):
        # Parse query parameters
        query_components = parse_qs(urlparse(self.path).query)
        file_name = query_components.get("file", [None])[0]
        
        if not file_name:
            self.send_response(400)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error_response = json.dumps({"error": "Template file name is required"})
            self.wfile.write(error_response.encode('utf-8'))
            return

        file_path = os.path.join(TEMPLATE_DIR, file_name)

        try:
            # Load the workbook using openpyxl
            workbook = load_workbook(file_path)

            # Dictionary to store placeholders by sheet name
            placeholders_by_sheet = {}

            # Iterate through each sheet in the workbook
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                placeholders_by_sheet[sheet] = set()

                # Iterate through each row and cell in the sheet
                for row in ws.iter_rows():
                    for cell in row:
                        value = str(cell.value)
                        if '{{' in value and '}}' in value:
                            matches = value.split('{{')[1:]
                            for match in matches:
                                placeholder = match.split('}}')[0].strip()
                                placeholders_by_sheet[sheet].add(placeholder)

            # Convert sets to lists for JSON serialization
            placeholders_by_sheet = {
                sheet: list(placeholders)
                for sheet, placeholders in placeholders_by_sheet.items()
            }

            # Return the placeholders as a JSON response
            response = json.dumps({"placeholders_by_sheet": placeholders_by_sheet})
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(response.encode('utf-8'))
        
        except Exception as e:
            print(f"Error extracting placeholders: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            error_response = json.dumps({"error": "Failed to extract placeholders from the template"})
            self.wfile.write(error_response.encode('utf-8'))
