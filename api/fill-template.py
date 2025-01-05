import os
import json
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook
from io import BytesIO

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../templates")
TEMPLATE_DIR = os.path.abspath(TEMPLATE_DIR)


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Parse incoming request body
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            request_data = json.loads(post_data.decode('utf-8'))

            file_name = request_data.get('fileName')
            form_data = request_data.get('formData')

            if not file_name or not form_data:
                self.send_error(400, "File name and form data are required")
                return

            file_path = os.path.join(TEMPLATE_DIR, file_name)

            if not os.path.exists(file_path):
                self.send_error(404, f"File not found: {file_path}")
                return

            # Load the workbook
            workbook = load_workbook(file_path)

            # Process workbook by replacing placeholders with form data
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and "{{" in cell.value and "}}" in cell.value:
                            placeholder = cell.value.strip("{{").strip("}}").strip()
                            if placeholder in form_data:
                                cell.value = form_data[placeholder]

            # Save the processed workbook to a buffer
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            # Prepare response for download
            self.send_response(200)
            self.send_header("Content-Disposition", f"attachment; filename=filled_{file_name}")
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.end_headers()

            # Write the buffer content to the response
            self.wfile.write(buffer.read())

        except Exception as e:
            print(f"Error processing request: {str(e)}")
            self.send_error(500, f"Internal server error: {str(e)}")
